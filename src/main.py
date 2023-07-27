import PySimpleGUI as sg
import common
import openpyxl
import numpy as np


def read_file(logger, filename):
    logger.log(f"Open File - {filename}")
    wb = openpyxl.load_workbook(filename=filename, data_only=True)

    first_col = 2
    first_row = 3
    for ws_name in wb.sheetnames:
        ws = wb.get_sheet_by_name(ws_name)
        num_cols = ws.max_column
        num_rows = len(ws['A'])

        all_cells = np.array([[cell.value for cell in row] for row in ws.iter_rows()])

        # all_cells is zero-indexed
        data = all_cells[(first_row - 1):(first_row - 1 + num_rows), (first_col - 1):(first_col - 1 + num_cols)]
        print(data)


def read_excel_into_array(file_paths):
    all_sheets_data = {}

    for file_path in file_paths:
        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames

        for sheet_name in sheet_names:
            if sheet_name != 'Help':
                sheet = workbook[sheet_name]
                data = []
                for row in sheet.iter_rows(values_only=True):
                    if row >= 2:
                        data.append(row)
                if sheet_name not in all_sheets_data:
                    all_sheets_data[sheet_name] = data
                else:
                    all_sheets_data[sheet_name].extend(data)

    return all_sheets_data


def determine_markup(logger, file1, file2, file3, file4, file5):
    logger.clear()
    file_paths = [file1, file2, file3, file4, file5]
    all_sheets_data = read_excel_into_array(file_paths)

    # Printing the data for each sheet
    for sheet_name, data in all_sheets_data.items():
        print(f"Sheet: {sheet_name}")
        for row in data:
            print(row)
        print("\n")


###################################################################################################################
# MAIN PROGRAM
###################################################################################################################
def main():
    sg.theme('DarkAmber')  # Add a touch of color

    # All the stuff inside your window.
    layout = [[sg.Text(
        "This program takes two group options files and compares them")
    ],
        [sg.Text("File 1: "), sg.Input(), sg.FileBrowse(key="-FILE1-")],
        [sg.Text("File 2: "), sg.Input(), sg.FileBrowse(key="-FILE2-")],
        [sg.Text("File 3: "), sg.Input(), sg.FileBrowse(key="-FILE3-")],
        [sg.Text("File 4: "), sg.Input(), sg.FileBrowse(key="-FILE4-")],
        [sg.Text("File 5: "), sg.Input(), sg.FileBrowse(key="-FILE5-")],
        [sg.Button("Submit")],
        [sg.Text("", key='-PROGRESS-')],
        [sg.Multiline(size=(70, 15), key='-TEXT-', autoscroll=True, expand_x=True, expand_y=True)]]

    allowed_extensions = ["xlsx", "xlsm", "xltx", "xltm"]

    # Create the Window
    window = sg.Window('Work out cost to sell markup in grids', layout, resizable=True)
    logger = common.Logger(window, "-PROGRESS-", "-TEXT-")
    # Event Loop to process "events" and get the "values" of the inputs
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == 'Cancel':  # if user closes window or clicks cancel
            break
        elif event == "Submit":
            if values["-FILE1-"] == '' or values["-FILE2-"] == '':
                logger.log('You must select the input files\n')
            elif values["-FILE1-"][-4:] not in allowed_extensions or \
                    values["-FILE2-"][-4:] not in allowed_extensions or \
                    values["-FILE3-"][-4:] not in allowed_extensions or \
                    values["-FILE4-"][-4:] not in allowed_extensions or \
                    values["-FILE5-"][-4:] not in allowed_extensions:
                logger.log(
                    'All files must be excel format - supported formats are: ' + ','.join(allowed_extensions) + '\n')
            else:
                determine_markup(logger, values["-FILE1-"], values["-FILE2-"], values["-FILE3-"],
                                 values["-FILE4-"], values["-FILE5-"])

    window.close()


if __name__ == '__main__':
    main()
